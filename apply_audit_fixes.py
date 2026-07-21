#!/usr/bin/env python3
"""
apply_audit_fixes.py — anvender rettelserne fra QA-multi-agent-auditten (2026-07-08).
Retter 11 bekræftede fejl + reelle advarsler, deduplikerer F24/Q8 (103 dubletter),
og kanoniserer by-/vejnavn-stavning mod DAWA. Genopbygger kort + xlsx til sidst.
Kør: python3 apply_audit_fixes.py
"""
import csv, os, re, math, json, urllib.request, urllib.parse, concurrent.futures
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font

OUT = os.path.dirname(os.path.abspath(__file__))
SCR = "/private/tmp/claude-501/-Users-sebastianpriess/8951f7bf-8429-48ee-b147-0aaeba151183/scratchpad"
UA = {'User-Agent': 'kartago-fix/1.0'}
PALETTE=['#4e79a7','#f28e2b','#59a14f','#e15759','#b07aa1','#edc948','#76b7b2','#ff9da7','#9c6b4f','#86bcb6','#d37295','#a0cbe8','#8cd17d','#b6992d','#499894','#e377c2','#1f77b4','#d62728','#2ca02c','#9467bd','#8c564b','#17becf','#bcbd22','#7f7f7f']
LOG=[]
def log(m): LOG.append(m); print(m)

def get(u):
    try: return json.loads(urllib.request.urlopen(urllib.request.Request(u,headers=UA),timeout=20).read())
    except Exception: return None
def geocode(q):
    d=get("https://api.dataforsyningen.dk/adgangsadresser?"+urllib.parse.urlencode({'q':q,'struktur':'mini','per_side':1}))
    if d: return round(float(d[0]['y']),6), round(float(d[0]['x']),6)
    return None,None
def rev(lat,lon):
    return get("https://api.dataforsyningen.dk/adgangsadresser/reverse?"+urllib.parse.urlencode({'x':lon,'y':lat,'struktur':'mini'}))
def loosenorm(s):
    s=(s or '').strip().lower().replace('æ','a').replace('ø','o').replace('å','a')
    s=s.replace('ae','a').replace('oe','o').replace('aa','a')
    return re.sub(r'[^a-z0-9]','',s)
def rd(fn):
    with open(os.path.join(OUT,fn),encoding='utf-8-sig') as f:
        r=list(csv.reader(f)); return r[0],r[1:]
def wr(fn,h,rows):
    with open(os.path.join(OUT,fn),'w',newline='',encoding='utf-8-sig') as f:
        w=csv.writer(f); w.writerow(h); w.writerows(rows)
def hav(a,b,c,d):
    R=6371000;r=math.pi/180;x=(c-a)*r;y=(d-b)*r
    return 2*R*math.asin(math.sqrt(math.sin(x/2)**2+math.cos(a*r)*math.cos(c*r)*math.sin(y/2)**2))
def one(rows,pred,label):
    m=[r for r in rows if pred(r)]
    if len(m)!=1: log(f"  !! {label}: forventede 1 række, fandt {len(m)}");
    return m[0] if m else None
def street_part(adr):  # del før sidste komma (vej+husnr)
    return adr.rsplit(',',1)[0].strip() if ',' in adr else adr.strip()
def rebuild(adr, postnr, by):
    return f"{street_part(adr)}, {postnr} {by}".strip()

h_ts,ts=rd('tankstationer_dk.csv'); h_sl,sl=rd('superladere_dk.csv'); h_ff,ff=rd('fastfood_kaeder_dk.csv')
log(f"Start: tank={len(ts)} super={len(sl)} fastfood={len(ff)}")

# ================= FASTFOOD (idx: 0 Kæde,1 Navn,2 Adr,3 Postnr,4 By,5 Lat,6 Lon) =================
r=one(ff,lambda r:r[0]=='Burger King' and 'Galten' in (r[1]+r[4]),"BK Galten")
if r: r[5],r[6]='56.14675','9.91837'; log(f"[1] BK Galten koord -> {r[5]},{r[6]}")
r=one(ff,lambda r:r[0]=='Burger King' and 'Esbjerg' in (r[1]+r[4]),"BK Esbjerg")
if r: r[5],r[6]='55.465456','8.458878'; log(f"[2] BK Esbjerg koord -> {r[5]},{r[6]}")
r=one(ff,lambda r:r[0]=='Gasoline Grill' and 'Tivoli' in r[1],"Gasoline Tivoli")
if r: r[3]='1620'; r[2]=rebuild(r[2],'1620',r[4]); log(f"[5] Gasoline Tivoli postnr 1630->1620")
r=one(ff,lambda r:r[0]=='Jagger' and 'Søborg' in (r[1]+r[4]),"Jagger Søborg")
if r:
    la,lo=geocode('Søborg Hovedgade 35, 2860 Søborg')
    r[3]='2860'; r[4]='Søborg'
    if la: r[5],r[6]=str(la),str(lo)
    r[2]=rebuild('Søborg Hovedgade 35','2860','Søborg')
    log(f"[7] Jagger Søborg postnr 2680->2860, koord -> {r[5]},{r[6]}")

# ================= SUPERLADERE (idx: 0 Op,1 Navn,2 Adr,3 Postnr,4 By,5 kW,6 Stik,7 Lat,8 Lon) =====
r=one(sl,lambda r:r[0]=='Circle K' and 'Frederiksborgvej' in r[1] and '2400' in r[3],"CircleK Frederiksborgvej")
if r: r[4]='København NV'; r[2]=rebuild(r[2],'2400','København NV'); log("[4] Circle K Frederiksborgvej By -> København NV")
r=one(sl,lambda r:r[0]=='E.ON' and 'Stilling' in r[1],"E.ON Stilling")
if r:
    r[3]='8660'; r[4]='Skanderborg'; r[2]='Ørstedsvej 4, 8660 Skanderborg'
    log("[6] E.ON Stilling -> Ørstedsvej 4, 8660 Skanderborg")
r=one(sl,lambda r:r[0]=='PowerGo',"PowerGo")
if r:
    r[4]='Nørre Aaby'; r[1]='PowerGo Nørre Aaby, Fynske Motorvej'
    r[2]=rebuild(r[2].replace('Norre Aaby','Nørre Aaby'),'5580','Nørre Aaby')
    log("[11] PowerGo Navn udfyldt + Nørre Aaby")
r=one(sl,lambda r:'Slagelse' in r[1] and ('Flextrafik' in r[1] or 'Sygehus' in r[1]),"Norlys Slagelse")
if r:
    la,lo=geocode('Fælledvej 1, 4200 Slagelse')
    if la: r[7],r[8]=str(la),str(lo); log(f"[15] Norlys Slagelse koord -> {la},{lo}")
r=one(sl,lambda r:r[0]=='Tesla' and 'Ikast' in r[1],"Tesla Ikast")
if r: r[2]='La Cours Vej 24, 7430 Ikast'; log("[19] Tesla Ikast adresse -> La Cours Vej 24")
r=one(sl,lambda r:r[0]=='Clever' and 'SPAR Fanø' in r[1],"SPAR Fanø")
if r:
    la,lo=geocode('Strandvejen 27, 6720 Fanø')
    r[2]='Strandvejen 27, 6720 Fanø'
    if la: r[7],r[8]=str(la),str(lo)
    log(f"[21] SPAR Fanø -> Strandvejen 27, koord {r[7]},{r[8]}")
# Circle K 'Copenhagen' engelsk bynavn
for r in sl:
    if r[0]=='Circle K' and r[4].strip().lower()=='copenhagen':
        r[4]='København S'; r[2]=rebuild(r[2],r[3],'København S'); log(f"[23] Circle K 'Copenhagen' -> København S ({r[1]})")
# Tesla Viby J (By Aarhus, postnr 8260)
for r in sl:
    if r[0]=='Tesla' and r[3]=='8260' and loosenorm(r[4])!=loosenorm('Viby J'):
        r[4]='Viby J'; r[2]=rebuild(r[2],'8260','Viby J'); log(f"[warn] Tesla {r[1]} By -> Viby J")
# E.ON malformet adresse (dobbelt postnr / 'Danmark' / parentes)
for r in sl:
    if r[0]=='E.ON' and (r[2].count(r[3])>1 or 'Danmark' in r[2] or re.search(r'\d{4}.*\d{4}',r[2])):
        a=rev(float(r[7]),float(r[8]))
        if a:
            r[3]=str(a['postnr']); r[4]=a['postnrnavn']
            r[2]=f"{a['vejnavn']} {a.get('husnr','')}".strip()+f", {a['postnr']} {a['postnrnavn']}"
            log(f"[25] E.ON malformet adresse renset -> {r[2]}")
# 1000 kW urimelig effekt
for r in sl:
    try:
        if float(r[5])>=1000:
            log(f"[FLAG] {r[0]} {r[1]} {r[2]} = {r[5]} kW (urimelig; sat til 400)")
            r[5]='400'
    except: pass

# ================= TANK (idx: 0 Mærke,1 Navn,2 Adr,3 Postnr,4 By,5 Lat,6 Lon) =================
# fjern kategori-fejl
before=len(ts)
ts=[r for r in ts if not ('BILLUND LUFTHAVN' in r[1].upper() and r[0]=='Circle K')]
log(f"[9] Circle K Billund Lufthavn fjernet ({before-len(ts)})")
before=len(ts)
ts=[r for r in ts if not ('Lystbådehavn' in r[1] or 'Lystbaadehavn' in r[1] or ('marina' in (r[1]+r[2]).lower()))]
log(f"[12] Marina/lystbådehavn fjernet ({before-len(ts)})")
# Dynamovej Herlev -> Søborg (begge mærker; Q8 overlever dedup)
for r in ts:
    if 'Dynamovej 2' in r[2] and r[3]=='2860':
        r[4]='Søborg'; r[1]=r[1].replace('Herlev','Søborg'); r[2]=rebuild(r[2],'2860','Søborg')
log("[8] Dynamovej 2, 2860: By Herlev -> Søborg")
# HK Benzin Faaborg koord
r=one(ts,lambda r:'HK' in r[0] and 'Faaborg' in r[1],"HK Benzin Faaborg")
if r:
    la,lo=geocode('Faaborgvej 49, 6818 Årre')
    if la: r[5],r[6]=str(la),str(lo); log(f"[13] HK Benzin Faaborg koord -> {la},{lo}")
# OIL! Ringe koord (lav præcision)
r=one(ts,lambda r:r[0]=='OIL!' and 'Ringe' in r[1],"OIL! Ringe")
if r:
    la,lo=geocode('Bygmestervej 1A, 5750 Ringe')
    if la: r[5],r[6]=str(la),str(lo); log(f"[14] OIL! Ringe koord -> {la},{lo}")
# OIL! Fjerritslev adresse
r=one(ts,lambda r:r[0]=='OIL!' and 'Fjerritslev' in r[1],"OIL! Fjerritslev")
if r: r[2]='Vestergaardsvej 46, 9690 Fjerritslev'; log("[18] OIL! Fjerritslev adresse -> Vestergaardsvej 46")
# Circle K EV Hillerød koord -> Herredsvejen 10
r=one(ts,lambda r:r[0]=='Circle K' and 'EV HILLER' in r[1].upper(),"CircleK EV Hillerød")
if r:
    la,lo=geocode('Herredsvejen 10, 3400 Hillerød')
    if la: r[5],r[6]=str(la),str(lo); log(f"[20] Circle K EV Hillerød koord -> {la},{lo}")
# Neksø -> Nexø
for r in ts:
    if loosenorm(r[4])=='nekso' and r[4]!='Nexø':
        r[4]='Nexø'; r[2]=rebuild(r[2],r[3],'Nexø'); log(f"[16] {r[1]} By Neksø -> Nexø")
# INGO Randers C -> NV (postnr 8920)
for r in ts:
    if r[3]=='8920' and 'RANDERS C' in r[4].upper():
        r[4]='Randers NV'; r[2]=rebuild(r[2],'8920','Randers NV'); log(f"[17] {r[1]} By Randers C -> Randers NV")
# OIL! Viby J akut-accent U+00B4 -> apostrof
for r in ts:
    if '´' in r[2]:
        r[2]=r[2].replace('´',"'"); log(f"[24] {r[1]} akut-accent -> apostrof: {r[2]}")

# ---- F24/Q8 dedup: fjern F24-rækker hvis koord matcher en Q8-række ----
q8coords=set((round(float(r[5]),6),round(float(r[6]),6)) for r in ts if r[0]=='Q8')
before=len(ts)
ts=[r for r in ts if not (r[0]=='F24' and (round(float(r[5]),6),round(float(r[6]),6)) in q8coords)]
log(f"[10] F24/Q8 dedup: fjernet {before-len(ts)} F24-dubletter")

# ---- '(ukendt)' navne i tank -> '<Mærke> <By>' ----
n=0
for r in ts:
    if r[1].strip().lower() in ('(ukendt)','ukendt',''):
        r[1]=f"{r[0]} {r[4]}".strip(); n+=1
log(f"[22] '(ukendt)'-navne udfyldt: {n}")

# ================= SYSTEMATISK STAVE-KANONISERING mod DAWA (by + vej) =================
def canonize(rows, latc, lonc, adrc, pnc, byc, name):
    coords=list({(round(float(r[latc]),6),round(float(r[lonc]),6)) for r in rows})
    cache={}
    def work(c):
        a=rev(c[0],c[1]); return c,a
    with concurrent.futures.ThreadPoolExecutor(max_workers=15) as ex:
        for c,a in ex.map(work, coords): cache[c]=a
    nby=nvej=0
    for r in rows:
        c=(round(float(r[latc]),6),round(float(r[lonc]),6)); a=cache.get(c)
        if not a: continue
        # by-stavning
        if a.get('postnrnavn') and loosenorm(a['postnrnavn'])==loosenorm(r[byc]) and r[byc]!=a['postnrnavn']:
            r[byc]=a['postnrnavn']; nby+=1
        # vejnavn-stavning i adresse-strengen
        sp=street_part(r[adrc]); m=re.match(r'^(.*?)(\s+\d+[A-Za-z]?)?$',sp)
        stname=(m.group(1) if m else sp).strip(); hus=(m.group(2) or '').strip()
        if a.get('vejnavn') and loosenorm(a['vejnavn'])==loosenorm(stname) and stname!=a['vejnavn']:
            newsp=f"{a['vejnavn']} {hus}".strip(); r[adrc]=rebuild(newsp+',', r[pnc], r[byc]); nvej+=1
        else:
            # sikr adresse<->postnr/by konsistens for rettede byer
            r[adrc]=rebuild(r[adrc], r[pnc], r[byc])
    log(f"  {name}: kanoniseret {nby} bynavne, {nvej} vejnavne (stavning/versaler mod DAWA)")

log("Kanoniserer stavning mod DAWA (kan tage et par minutter) ...")
canonize(sl, 7,8,2,3,4, "superladere")
canonize(ts, 5,6,2,3,4, "tankstationer")
canonize(ff, 5,6,2,3,4, "fastfood")

# ================= SORTÉR + GEM =================
ts.sort(key=lambda x:(x[0],str(x[3]))); sl.sort(key=lambda x:(x[0],str(x[3]))); ff.sort(key=lambda x:(x[0],str(x[3])))
wr('tankstationer_dk.csv',h_ts,ts); wr('superladere_dk.csv',h_sl,sl); wr('fastfood_kaeder_dk.csv',h_ff,ff)
log(f"\nGEMT: tank={len(ts)} super={len(sl)} fastfood={len(ff)}")
log("Tank-mærker: "+str(dict(Counter(r[0] for r in ts).most_common())))

# ================= GENOPBYG kort + xlsx =================
def num(v):
    try:return float(v)
    except:return None
def build(rws,bi_,ni,ai,li,gi,extra):
    cnt=Counter(r[bi_] for r in rws); ordered=[b for b,_ in cnt.most_common()]
    brands=[{'n':b,'c':PALETTE[i%len(PALETTE)],'k':cnt[b]} for i,b in enumerate(ordered)]
    idx={b['n']:i for i,b in enumerate(brands)}; pts=[]
    for r in rws:
        la=num(r[li]);lo=num(r[gi])
        if la is None or lo is None: continue
        pts.append([round(la,5),round(lo,5),idx[r[bi_]],r[ni],r[ai],extra(r)])
    return {'brands':brands,'pts':pts}
charge=build(sl,0,1,2,7,8,lambda r:" · ".join([x for x in [(str(r[5])+" kW" if r[5] else ""),r[6]] if x]))
food=build(ff,0,1,2,5,6,lambda r:""); tank=build(ts,0,1,2,5,6,lambda r:"")
DATA={'charge':{'label':'Superladere','sub':'≥250 kW · Tesla-niveau','total':len(charge['pts']),**charge},
      'food':{'label':'Fastfood','sub':'Fastfood-kæder','total':len(food['pts']),**food},
      'tank':{'label':'Tankstationer','sub':'Alle mærker','total':len(tank['pts']),**tank}}
dj=json.dumps(DATA,ensure_ascii=False,separators=(',',':'))
open(os.path.join(OUT,'kort.html'),'w',encoding='utf-8').write(open(SCR+'/map_template.html',encoding='utf-8').read().replace('/*__DATA__*/',dj).replace('/*__OUTLINE__*/',json.dumps(json.load(open(SCR+'/dk_outline.json')),ensure_ascii=False,separators=(',',':'))))
open(os.path.join(OUT,'kort_soeg.html'),'w',encoding='utf-8').write(open(SCR+'/map_local_template.html',encoding='utf-8').read().replace('/*__DATA__*/',dj))
wb=Workbook()
def fill(ws,head,data,widths):
    ws.append(head)
    for c in ws[1]: c.font=Font(bold=True)
    for r in data: ws.append(r)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for i,wd in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=wd
ws=wb.active; ws.title="Superladere"; fill(ws,h_sl,sl,[16,26,42,8,16,9,10,11,11])
fill(wb.create_sheet("Fastfood"),h_ff,ff,[18,28,44,8,16,11,11])
fill(wb.create_sheet("Tankstationer"),h_ts,ts,[14,26,42,8,16,11,11])
wb.save(os.path.join(OUT,'kaede_adresser.xlsx'))
log("Genopbygget kort.html, kort_soeg.html, kaede_adresser.xlsx")
open(os.path.join(OUT,'audit_fixes_log.txt'),'w',encoding='utf-8').write("\n".join(LOG))
