#!/usr/bin/env python3
"""
rebuild.py — genopbyg kort_soeg.html + kaede_adresser.xlsx ud fra de tre CSV'er.
Kør: python3 rebuild.py   (efter refresh_data.py / manuelle CSV-rettelser)

Kortet er sin egen skabelon: DATA-blokken (const DATA = {...};) i kort_soeg.html
udskiftes in-place, alt andet i filen bevares byte-for-byte. Excel bygges fra bunden.
"""
import csv, os, json, re, datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font

OUT = os.path.dirname(os.path.abspath(__file__))
PALETTE = ['#4e79a7','#f28e2b','#59a14f','#e15759','#b07aa1','#edc948','#76b7b2','#ff9da7',
           '#9c6b4f','#86bcb6','#d37295','#a0cbe8','#8cd17d','#b6992d','#499894','#e377c2',
           '#1f77b4','#d62728','#2ca02c','#9467bd','#8c564b','#17becf','#bcbd22','#7f7f7f']

def rd(fn):
    with open(os.path.join(OUT, fn), encoding='utf-8-sig') as f:
        r = list(csv.reader(f)); return r[0], r[1:]

def num(v):
    try: return float(v)
    except: return None

def build(rows, bi, ni, ai, li, gi, extra):
    """bi=mærke, ni=navn, ai=adresse, li=lat, gi=lon, extra=fn(row)->str"""
    cnt = Counter(r[bi] for r in rows)
    ordered = [b for b, _ in cnt.most_common()]
    brands = [{'n': b, 'c': PALETTE[i % len(PALETTE)], 'k': cnt[b]} for i, b in enumerate(ordered)]
    idx = {b['n']: i for i, b in enumerate(brands)}
    pts = []
    for r in rows:
        la, lo = num(r[li]), num(r[gi])
        if la is None or lo is None: continue
        pts.append([round(la, 5), round(lo, 5), idx[r[bi]], r[ni], r[ai], extra(r)])
    return {'brands': brands, 'pts': pts}

# ---- læs CSV'er (nuværende skema) ----
# superladere: 0 Operatør,1 Navn,2 Adresse,3 Postnr,4 By,5 Effekt_kW,6 Stik,7 Antal_ladere,8 Lat,9 Lon
# fastfood/tank: 0 Mærke,1 Navn,2 Adresse,3 Postnr,4 By,5 Lat,6 Lon
h_sl, sl = rd('superladere_dk.csv')
h_ff, ff = rd('fastfood_kaeder_dk.csv')
h_ts, ts = rd('tankstationer_dk.csv')

def charge_extra(r):
    parts = [f"{r[5]} kW" if r[5] else "", r[6], f"{r[7]} ladestandere" if r[7] else ""]
    return " · ".join([x for x in parts if x])

charge = build(sl, 0, 1, 2, 8, 9, charge_extra)
food   = build(ff, 0, 1, 2, 5, 6, lambda r: "")
tank   = build(ts, 0, 1, 2, 5, 6, lambda r: "")
DATA = {'charge': {'label': 'Superladere',   'sub': '≥250 kW · Tesla-niveau', 'total': len(charge['pts']), **charge},
        'food':   {'label': 'Fastfood',      'sub': 'Fastfood-kæder',        'total': len(food['pts']),   **food},
        'tank':   {'label': 'Tankstationer', 'sub': 'Alle mærker',           'total': len(tank['pts']),   **tank}}
dj = json.dumps(DATA, ensure_ascii=False, separators=(',', ':'))

# ---- swap DATA-blokken i kort_soeg.html (bevar alt andet) ----
def swap_data(html, new_json):
    key = 'const DATA = '
    i = html.find(key)
    if i < 0: raise SystemExit("kunne ikke finde 'const DATA =' i kort_soeg.html")
    i += len(key)
    depth = 0; j = i; instr = False; esc = False
    while j < len(html):
        ch = html[j]
        if instr:
            if esc: esc = False
            elif ch == '\\': esc = True
            elif ch == '"': instr = False
        else:
            if ch == '"': instr = True
            elif ch == '{': depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0: j += 1; break
        j += 1
    return html[:i] + new_json + html[j:]

MONTHS_DA = ['januar','februar','marts','april','maj','juni','juli','august',
             'september','oktober','november','december']
def stamp_today():
    d = datetime.date.today()
    return f"Data pr. {d.day}. {MONTHS_DA[d.month - 1]} {d.year}"

hp = os.path.join(OUT, 'kort_soeg.html')
html = open(hp, encoding='utf-8').read()
html = swap_data(html, dj)
# opdatér synligt datostempel (elementet bevares mellem builds; matcher intet hvis fjernet)
html = re.sub(r'(<p class="stamp" id="datastamp"[^>]*>).*?(</p>)',
              lambda m: m.group(1) + stamp_today() + m.group(2), html, count=1)
open(hp, 'w', encoding='utf-8').write(html)

# ---- data-feed (til "data-feed + fallback"): samme DATA + datostempel ----
# Kortet henter denne fil hvis FEED_URL er sat i kort_soeg.html; ellers bruges den indbyggede DATA.
feed = {'stamp': stamp_today(), **DATA}
open(os.path.join(OUT, 'retailkort_data.json'), 'w', encoding='utf-8').write(
    json.dumps(feed, ensure_ascii=False, separators=(',', ':')))

# ---- genopbyg Excel ----
def to_num(v):
    """tal-kolonner skrives som tal (float m. decimal, ellers int); tomt -> blank; ellers tekst."""
    if v is None or v == '': return None
    try:
        return float(v) if '.' in v else int(v)
    except (ValueError, TypeError):
        return v

def fill(ws, head, data, widths, numcols=()):
    ws.append(head)
    for c in ws[1]: c.font = Font(bold=True)
    ns = set(numcols)
    for r in data:
        ws.append([to_num(v) if i in ns else v for i, v in enumerate(r)])
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = wd

wb = Workbook()
ws = wb.active; ws.title = "Superladere"
# numcols: Effekt_kW(5), Antal_ladere(7), Latitude(8), Longitude(9) — Postnr holdes som tekst (ID)
fill(ws, h_sl, sl, [16, 26, 42, 8, 15, 9, 10, 13, 11, 11], numcols=(5, 7, 8, 9))
fill(wb.create_sheet("Fastfood"), h_ff, ff, [18, 28, 44, 8, 16, 11, 11], numcols=(5, 6))
fill(wb.create_sheet("Tankstationer"), h_ts, ts, [14, 26, 42, 8, 16, 11, 11], numcols=(5, 6))
wb.save(os.path.join(OUT, 'kaede_adresser.xlsx'))

print(f"Genopbygget: kort_soeg.html ({len(charge['pts'])} superladere, "
      f"{len(food['pts'])} fastfood, {len(tank['pts'])} tank) + kaede_adresser.xlsx "
      f"+ retailkort_data.json (feed)")
